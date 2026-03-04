"""
=====================================================
  STOCK ANALYZER WEB APP v4.0 — app.py
=====================================================
  NEW FEATURES:
  - Compare 2 stocks side by side
  - Portfolio tracker
  - News feed
  - Candlestick chart (OHLC)
  - Excel export
  - Watchlist
  - Prominent Buy/Sell/Hold signal
"""

from flask import Flask, request, jsonify, send_file
import yfinance as yf
import pandas as pd
import numpy as np
import io
import os

app = Flask(__name__)

# ── Serve HTML ────────────────────────────────────
@app.route("/")
def home():
    try:
        with open("index.html", "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return "<h2>Error: index.html not found!</h2>", 404

# ── Helpers ──────────────────────────────────────
def sf(val, digits=2):
    try:
        v = float(val)
        if pd.isna(v): return None
        return round(v, digits)
    except: return None

def fmt_mc(mc):
    if not mc: return "N/A"
    if mc >= 1e12: return f"{mc/1e12:.2f}T"
    if mc >= 1e9:  return f"{mc/1e9:.2f}B"
    if mc >= 1e6:  return f"{mc/1e6:.2f}M"
    return str(mc)

def calc_rsi(prices, period=14):
    delta    = prices.diff()
    gains    = delta.where(delta > 0, 0)
    losses   = -delta.where(delta < 0, 0)
    avg_gain = gains.rolling(window=period).mean()
    avg_loss = losses.rolling(window=period).mean()
    rs       = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))

def calc_macd(prices, fast=12, slow=26, signal=9):
    ema_fast    = prices.ewm(span=fast,   adjust=False).mean()
    ema_slow    = prices.ewm(span=slow,   adjust=False).mean()
    macd_line   = ema_fast - ema_slow
    signal_line = macd_line.ewm(span=signal, adjust=False).mean()
    histogram   = macd_line - signal_line
    return macd_line, signal_line, histogram

def calc_bb(prices, period=20, num_std=2):
    middle = prices.rolling(window=period).mean()
    std    = prices.rolling(window=period).std()
    upper  = middle + (num_std * std)
    lower  = middle - (num_std * std)
    bwidth = (upper - lower) / middle * 100
    return upper, middle, lower, bwidth

def calc_stochastic(high, low, close, k_period=14, d_period=3):
    lowest_low   = low.rolling(window=k_period).min()
    highest_high = high.rolling(window=k_period).max()
    k = 100 * (close - lowest_low) / (highest_high - lowest_low)
    d = k.rolling(window=d_period).mean()
    return k, d

def calc_atr(high, low, close, period=14):
    high_low   = high - low
    high_close = (high - close.shift()).abs()
    low_close  = (low  - close.shift()).abs()
    true_range = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    return true_range.rolling(window=period).mean()

def calc_fibonacci(high_price, low_price):
    diff = high_price - low_price
    return {
        "0% (High)":  round(high_price, 2),
        "23.6%":      round(high_price - 0.236 * diff, 2),
        "38.2%":      round(high_price - 0.382 * diff, 2),
        "50.0%":      round(high_price - 0.500 * diff, 2),
        "61.8%":      round(high_price - 0.618 * diff, 2),
        "78.6%":      round(high_price - 0.786 * diff, 2),
        "100% (Low)": round(low_price,  2),
    }

def calc_support_resistance(close, high, low, n_levels=3):
    window = 5
    resistance_levels = []
    support_levels    = []
    for i in range(window, len(close) - window):
        if high.iloc[i] == high.iloc[i-window:i+window].max():
            resistance_levels.append(float(high.iloc[i]))
        if low.iloc[i] == low.iloc[i-window:i+window].min():
            support_levels.append(float(low.iloc[i]))
    def cluster(levels, tol=0.02):
        if not levels: return []
        levels = sorted(set(levels))
        result = [levels[0]]
        for lvl in levels[1:]:
            if (lvl - result[-1]) / result[-1] > tol:
                result.append(lvl)
        return result
    resistance_levels = cluster(resistance_levels)
    support_levels    = cluster(support_levels)
    latest = float(close.iloc[-1])
    res = sorted([r for r in resistance_levels if r > latest])[:n_levels]
    sup = sorted([s for s in support_levels    if s < latest], reverse=True)[:n_levels]
    return sup, res

def full_analysis(symbol, period="3mo"):
    """Full analysis for a stock — used by both /analyze and /compare"""
    if period == "1d":
        data = yf.download(symbol, period="1d",  interval="5m",  progress=False)
    elif period == "5d":
        data = yf.download(symbol, period="5d",  interval="15m", progress=False)
    elif period == "1mo":
        data = yf.download(symbol, period="1mo", interval="1d",  progress=False)
    elif period == "1y":
        data = yf.download(symbol, period="1y",  interval="1d",  progress=False)
    else:
        data = yf.download(symbol, period="3mo", interval="1d",  progress=False)

    if data.empty: return None

    stock = yf.Ticker(symbol)
    info  = stock.info

    close  = data["Close"].squeeze()
    high   = data["High"].squeeze()
    low    = data["Low"].squeeze()
    volume = data["Volume"].squeeze()
    open_  = data["Open"].squeeze()

    latest = float(close.iloc[-1])
    first  = float(close.iloc[0])
    change = latest - first
    pct    = (change / first) * 100
    daily_ret     = close.pct_change() * 100
    avg_daily     = float(daily_ret.mean())

    company_name  = info.get("longName", symbol)
    currency      = info.get("currency", "USD")
    sector        = info.get("sector",   "N/A")
    industry      = info.get("industry", "N/A")
    market_cap    = info.get("marketCap", None)
    pe_ratio      = info.get("trailingPE", None)
    pb_ratio      = info.get("priceToBook", None)
    eps           = info.get("trailingEps", None)
    div_yield     = info.get("dividendYield", None)
    week52_high   = info.get("fiftyTwoWeekHigh", None)
    week52_low    = info.get("fiftyTwoWeekLow",  None)
    description   = info.get("longBusinessSummary", "")
    employees     = info.get("fullTimeEmployees", None)
    country       = info.get("country", "N/A")
    website       = info.get("website", "N/A")
    roe           = info.get("returnOnEquity", None)
    revenue       = info.get("totalRevenue", None)
    profit_margin = info.get("profitMargins", None)
    debt_equity   = info.get("debtToEquity", None)
    current_ratio = info.get("currentRatio", None)

    fundamentals = {
        "company":       company_name,
        "sector":        sector,
        "industry":      industry,
        "country":       country,
        "employees":     f"{employees:,}" if employees else "N/A",
        "website":       website,
        "description":   description[:500] + "..." if len(description) > 500 else description,
        "market_cap":    fmt_mc(market_cap),
        "pe_ratio":      sf(pe_ratio)   if pe_ratio   else "N/A",
        "pb_ratio":      sf(pb_ratio)   if pb_ratio   else "N/A",
        "eps":           sf(eps)        if eps        else "N/A",
        "dividend_yield":f"{sf(div_yield*100)}%" if div_yield else "N/A",
        "week52_high":   sf(week52_high) if week52_high else "N/A",
        "week52_low":    sf(week52_low)  if week52_low  else "N/A",
        "roe":           f"{sf(roe*100)}%" if roe else "N/A",
        "revenue":       fmt_mc(revenue),
        "profit_margin": f"{sf(profit_margin*100)}%" if profit_margin else "N/A",
        "debt_equity":   sf(debt_equity)   if debt_equity   else "N/A",
        "current_ratio": sf(current_ratio) if current_ratio else "N/A",
    }

    sma7  = sf(close.rolling(7).mean().iloc[-1])  if len(data) >= 7  else None
    sma20 = sf(close.rolling(20).mean().iloc[-1]) if len(data) >= 20 else None

    rsi_signal = "UNKNOWN"; current_rsi = None; rsi_series = None
    if len(data) >= 14:
        rsi_series  = calc_rsi(close)
        current_rsi = sf(rsi_series.iloc[-1])
        if current_rsi > 70:   rsi_signal = "OVERBOUGHT"
        elif current_rsi < 30: rsi_signal = "OVERSOLD"
        else:                  rsi_signal = "NORMAL"

    macd_signal = "UNKNOWN"
    macd_line = signal_line = histogram = None
    curr_macd = curr_sig = curr_hist = None
    if len(data) >= 26:
        macd_line, signal_line, histogram = calc_macd(close)
        curr_macd = sf(macd_line.iloc[-1], 4)
        curr_sig  = sf(signal_line.iloc[-1], 4)
        curr_hist = sf(histogram.iloc[-1], 4)
        prev_macd = float(macd_line.iloc[-2])
        prev_sig  = float(signal_line.iloc[-2])
        bull_cross = (prev_macd < prev_sig) and (float(macd_line.iloc[-1]) > float(signal_line.iloc[-1]))
        bear_cross = (prev_macd > prev_sig) and (float(macd_line.iloc[-1]) < float(signal_line.iloc[-1]))
        if bull_cross:   macd_signal = "BULLISH_CROSSOVER"
        elif bear_cross: macd_signal = "BEARISH_CROSSOVER"
        elif float(macd_line.iloc[-1]) > float(signal_line.iloc[-1]): macd_signal = "BULLISH"
        else: macd_signal = "BEARISH"

    bb_signal = "UNKNOWN"; bb_squeeze = False
    bb_upper_s = bb_lower_s = bb_bw_s = None
    curr_upper = curr_lower = curr_bw = avg_bw = None
    if len(data) >= 20:
        bbu, bbm, bbl, bbbw = calc_bb(close)
        curr_upper = sf(bbu.iloc[-1]); curr_lower = sf(bbl.iloc[-1])
        curr_bw    = sf(bbbw.iloc[-1]); avg_bw = sf(float(bbbw.mean()))
        bb_upper_s = [sf(v) for v in bbu.values]
        bb_lower_s = [sf(v) for v in bbl.values]
        bb_bw_s    = [sf(v) for v in bbbw.values]
        if latest >= float(bbu.iloc[-1]):   bb_signal = "ABOVE_UPPER"
        elif latest <= float(bbl.iloc[-1]): bb_signal = "BELOW_LOWER"
        elif latest > float(bbm.iloc[-1]):  bb_signal = "UPPER_ZONE"
        else:                               bb_signal = "LOWER_ZONE"
        if float(bbbw.iloc[-1]) < float(bbbw.mean()) * 0.7: bb_squeeze = True

    stoch_signal = "UNKNOWN"; stoch_k_s = stoch_d_s = None; curr_k = curr_d = None
    if len(data) >= 14:
        k, d = calc_stochastic(high, low, close)
        curr_k = sf(k.iloc[-1]); curr_d = sf(d.iloc[-1])
        stoch_k_s = [sf(v) for v in k.values]
        stoch_d_s = [sf(v) for v in d.values]
        prev_k = float(k.iloc[-2]); prev_d = float(d.iloc[-2])
        bull_cs = (prev_k < prev_d) and (float(k.iloc[-1]) > float(d.iloc[-1]))
        bear_cs = (prev_k > prev_d) and (float(k.iloc[-1]) < float(d.iloc[-1]))
        if curr_k > 80:    stoch_signal = "OVERBOUGHT"
        elif curr_k < 20:  stoch_signal = "OVERSOLD"
        elif bull_cs:      stoch_signal = "BULLISH_CROSS"
        elif bear_cs:      stoch_signal = "BEARISH_CROSS"
        else:              stoch_signal = "NORMAL"

    curr_atr = atr_pct = None
    if len(data) >= 14:
        atr_series = calc_atr(high, low, close)
        curr_atr   = sf(atr_series.iloc[-1])
        atr_pct    = sf((curr_atr / latest) * 100) if curr_atr else None

    fib_levels  = calc_fibonacci(float(high.max()), float(low.min()))
    supports, resistances = [], []
    if len(data) >= 15:
        supports, resistances = calc_support_resistance(close, high, low)

    avg_vol_val  = float(volume.mean())
    latest_vol   = float(volume.iloc[-1])
    vol_ratio    = latest_vol / avg_vol_val
    unusual_days = int((volume > avg_vol_val * 2).sum())
    if vol_ratio >= 3:     vol_signal = "EXTREMELY_HIGH"
    elif vol_ratio >= 2:   vol_signal = "HIGH"
    elif vol_ratio >= 1.5: vol_signal = "ABOVE_AVERAGE"
    elif vol_ratio < 0.5:  vol_signal = "LOW"
    else:                  vol_signal = "NORMAL"

    # Score
    score = 0; breakdown = []
    def add(cond, pos_text, neg_text, pos_val, neg_val):
        nonlocal score
        if cond: score += pos_val; breakdown.append({"text":pos_text,"val":f"+{pos_val}","pos":True})
        else:    score += neg_val; breakdown.append({"text":neg_text,"val":str(neg_val),"pos":False})

    add(sma20 and latest > sma20, "Price ABOVE 20-day average","Price BELOW 20-day average",1,-1)
    add(sma7 and sma20 and sma7>sma20,"Short-term trend UP","Short-term trend DOWN",1,-1)
    add(pct>0, f"Stock gained {abs(pct):.1f}%",f"Stock lost {abs(pct):.1f}%",1,-1)

    if rsi_signal=="OVERSOLD":    score+=2; breakdown.append({"text":f"RSI OVERSOLD ({current_rsi}) bounce likely!","val":"+2","pos":True})
    elif rsi_signal=="OVERBOUGHT":score-=2; breakdown.append({"text":f"RSI OVERBOUGHT ({current_rsi}) may drop","val":"-2","pos":False})
    elif rsi_signal=="NORMAL":    breakdown.append({"text":f"RSI NORMAL ({current_rsi})","val":"0","pos":None})

    if macd_signal=="BULLISH_CROSSOVER":  score+=3; breakdown.append({"text":"MACD BULLISH CROSSOVER!","val":"+3","pos":True})
    elif macd_signal=="BEARISH_CROSSOVER":score-=3; breakdown.append({"text":"MACD BEARISH CROSSOVER!","val":"-3","pos":False})
    elif macd_signal=="BULLISH":          score+=1; breakdown.append({"text":"MACD Bullish momentum","val":"+1","pos":True})
    elif macd_signal=="BEARISH":          score-=1; breakdown.append({"text":"MACD Bearish momentum","val":"-1","pos":False})

    if bb_signal=="BELOW_LOWER":  score+=2; breakdown.append({"text":"BB Price at lower band","val":"+2","pos":True})
    elif bb_signal=="ABOVE_UPPER":score-=2; breakdown.append({"text":"BB Price at upper band","val":"-2","pos":False})
    elif bb_signal=="UPPER_ZONE": score+=1; breakdown.append({"text":"BB Upper zone","val":"+1","pos":True})
    elif bb_signal=="LOWER_ZONE": score-=1; breakdown.append({"text":"BB Lower zone","val":"-1","pos":False})
    if bb_squeeze: breakdown.append({"text":"BB SQUEEZE - Big move coming!","val":"!","pos":None})

    if stoch_signal=="OVERSOLD":    score+=2; breakdown.append({"text":f"STOCH OVERSOLD ({curr_k})","val":"+2","pos":True})
    elif stoch_signal=="OVERBOUGHT":score-=2; breakdown.append({"text":f"STOCH OVERBOUGHT ({curr_k})","val":"-2","pos":False})
    elif stoch_signal=="BULLISH_CROSS":score+=1; breakdown.append({"text":"STOCH Bullish cross","val":"+1","pos":True})
    elif stoch_signal=="BEARISH_CROSS":score-=1; breakdown.append({"text":"STOCH Bearish cross","val":"-1","pos":False})

    if vol_signal in ["HIGH","EXTREMELY_HIGH"]:
        if pct>=0: score+=2; breakdown.append({"text":"HIGH VOLUME + Price UP","val":"+2","pos":True})
        else:      score-=2; breakdown.append({"text":"HIGH VOLUME + Price DOWN","val":"-2","pos":False})

    if score >= 6:    rec = "VERY STRONG BUY"
    elif score >= 4:  rec = "STRONG BUY"
    elif score >= 2:  rec = "BUY"
    elif score >= 0:  rec = "HOLD"
    elif score >= -2: rec = "WEAK SELL"
    elif score >= -4: rec = "SELL"
    else:             rec = "STRONG SELL"

    # OHLC data for candlestick
    dates  = [str(d)[:16] for d in close.index]
    ohlc   = [{"t": str(close.index[i])[:16],
               "o": sf(float(open_.iloc[i])),
               "h": sf(float(high.iloc[i])),
               "l": sf(float(low.iloc[i])),
               "c": sf(float(close.iloc[i]))} for i in range(len(close))]

    return {
        "company":        company_name,
        "symbol":         symbol,
        "currency":       currency,
        "latest":         sf(latest),
        "change":         sf(change),
        "pct_change":     sf(pct),
        "avg_price":      sf(float(close.mean())),
        "highest":        sf(float(high.max())),
        "lowest":         sf(float(low.min())),
        "best_day":       sf(float(daily_ret.max())),
        "worst_day":      sf(float(daily_ret.min())),
        "avg_daily":      sf(avg_daily, 3),
        "sma7":           sma7,
        "sma20":          sma20,
        "rsi":            current_rsi,
        "rsi_signal":     rsi_signal,
        "macd":           curr_macd,
        "macd_sig_line":  curr_sig,
        "macd_hist_val":  curr_hist,
        "macd_signal":    macd_signal,
        "bb_upper":       curr_upper,
        "bb_lower":       curr_lower,
        "bb_bandwidth":   curr_bw,
        "bb_avg_bw":      avg_bw,
        "bb_signal":      bb_signal,
        "bb_squeeze":     bb_squeeze,
        "stoch_k":        curr_k,
        "stoch_d":        curr_d,
        "stoch_signal":   stoch_signal,
        "atr":            curr_atr,
        "atr_pct":        atr_pct,
        "fib_levels":     fib_levels,
        "supports":       [round(s,2) for s in supports],
        "resistances":    [round(r,2) for r in resistances],
        "avg_volume":     round(avg_vol_val),
        "latest_volume":  round(latest_vol),
        "vol_ratio":      sf(vol_ratio),
        "vol_signal":     vol_signal,
        "unusual_days":   unusual_days,
        "score":          score,
        "recommendation": rec,
        "breakdown":      breakdown,
        "fundamentals":   fundamentals,
        "chart": {
            "dates":    dates,
            "ohlc":     ohlc,
            "prices":   [sf(v) for v in close.values],
            "volumes":  [int(v) for v in volume.values],
            "sma7":     [sf(v) for v in close.rolling(7).mean().values],
            "sma20":    [sf(v) for v in close.rolling(20).mean().values],
            "bb_upper": bb_upper_s,
            "bb_lower": bb_lower_s,
            "rsi":      [sf(v) for v in rsi_series.values] if rsi_series is not None else None,
            "macd":     [sf(v,4) for v in macd_line.values]   if macd_line   is not None else None,
            "macd_sig": [sf(v,4) for v in signal_line.values] if signal_line is not None else None,
            "macd_hist":[sf(v,4) for v in histogram.values]   if histogram   is not None else None,
            "bb_bw":    bb_bw_s,
            "stoch_k":  stoch_k_s,
            "stoch_d":  stoch_d_s,
        }
    }


# ── Analyze endpoint ──────────────────────────────
@app.route("/analyze", methods=["POST"])
def analyze():
    try:
        body   = request.get_json()
        symbol = body.get("symbol","").upper().strip()
        period = body.get("period","3mo")
        if not symbol:
            return jsonify({"error":"Please enter a stock symbol!"}), 400
        result = full_analysis(symbol, period)
        if result is None:
            return jsonify({"error":f"No data found for '{symbol}'"}), 404
        return jsonify(result)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ── Compare 2 stocks ──────────────────────────────
@app.route("/compare", methods=["POST"])
def compare():
    try:
        body    = request.get_json()
        symbol1 = body.get("symbol1","").upper().strip()
        symbol2 = body.get("symbol2","").upper().strip()
        period  = body.get("period","3mo")
        if not symbol1 or not symbol2:
            return jsonify({"error":"Please enter both stock symbols!"}), 400
        r1 = full_analysis(symbol1, period)
        r2 = full_analysis(symbol2, period)
        if r1 is None: return jsonify({"error":f"No data for '{symbol1}'"}), 404
        if r2 is None: return jsonify({"error":f"No data for '{symbol2}'"}), 404
        return jsonify({"stock1": r1, "stock2": r2})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ── News endpoint ─────────────────────────────────
@app.route("/news", methods=["POST"])
def get_news():
    try:
        body   = request.get_json()
        symbol = body.get("symbol","").upper().strip()
        if not symbol:
            return jsonify({"error":"Please enter a stock symbol!"}), 400
        stock = yf.Ticker(symbol)
        news  = stock.news or []
        result = []
        for item in news[:15]:
            result.append({
                "title":     item.get("title","No title"),
                "publisher": item.get("publisher","Unknown"),
                "link":      item.get("link","#"),
                "time":      item.get("providerPublishTime", 0),
            })
        return jsonify({"news": result, "symbol": symbol})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Excel Export ──────────────────────────────────
@app.route("/export", methods=["POST"])
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        body   = request.get_json()
        symbol = body.get("symbol","").upper().strip()
        period = body.get("period","3mo")

        if not symbol:
            return jsonify({"error":"Please enter a symbol"}), 400

        d = full_analysis(symbol, period)
        if d is None:
            return jsonify({"error":f"No data for '{symbol}'"}), 404

        wb = openpyxl.Workbook()

        # Colors
        dark_fill   = PatternFill("solid", fgColor="0D1117")
        green_fill  = PatternFill("solid", fgColor="00E676")
        red_fill    = PatternFill("solid", fgColor="FF1744")
        blue_fill   = PatternFill("solid", fgColor="2979FF")
        header_fill = PatternFill("solid", fgColor="161B22")
        white_font  = Font(color="FFFFFF", bold=True, name="Consolas")
        black_font  = Font(color="000000", bold=True, name="Consolas")
        muted_font  = Font(color="8B949E", name="Consolas")
        green_font  = Font(color="00E676", bold=True, name="Consolas")
        red_font    = Font(color="FF1744", bold=True, name="Consolas")
        yellow_font = Font(color="FFEA00", bold=True, name="Consolas")

        def hdr(ws, row, col, text):
            c = ws.cell(row=row, column=col, value=text)
            c.fill = header_fill; c.font = Font(color="4A5568", name="Consolas", size=9)
            c.alignment = Alignment(horizontal="left")
            return c

        def val(ws, row, col, text, color="white"):
            c = ws.cell(row=row, column=col, value=text)
            c.fill = dark_fill
            if color == "green": c.font = green_font
            elif color == "red": c.font = red_font
            elif color == "yellow": c.font = yellow_font
            elif color == "muted": c.font = muted_font
            else: c.font = Font(color="CDD9E5", name="Consolas")
            return c

        # ── Sheet 1: Summary ──
        ws1 = wb.active; ws1.title = "Summary"
        ws1.sheet_view.showGridLines = False
        for row in ws1.iter_rows(min_row=1, max_row=50, min_col=1, max_col=6):
            for cell in row: cell.fill = dark_fill

        ws1.column_dimensions["A"].width = 22
        ws1.column_dimensions["B"].width = 20
        ws1.column_dimensions["C"].width = 22
        ws1.column_dimensions["D"].width = 20

        # Title
        t = ws1.cell(row=1, column=1, value=f"STOCK ANALYZER — {d['symbol']}")
        t.font = Font(color="00E676", bold=True, name="Consolas", size=14)
        t.fill = dark_fill

        ws1.cell(row=2, column=1, value=d["company"]).font = Font(color="FFFFFF", name="Consolas", size=11)
        ws1.cell(row=2, column=1).fill = dark_fill

        # Signal box
        rec_color = "green" if d["score"]>=2 else "red" if d["score"]<=-2 else "yellow"
        ws1.cell(row=1, column=3, value="SIGNAL").font = Font(color="4A5568", name="Consolas", size=9)
        ws1.cell(row=1, column=3).fill = dark_fill
        val(ws1, 2, 3, d["recommendation"], rec_color)

        ws1.cell(row=1, column=4, value="SCORE").font = Font(color="4A5568", name="Consolas", size=9)
        ws1.cell(row=1, column=4).fill = dark_fill
        val(ws1, 2, 4, f"{d['score']} / 14")

        # Price info
        r = 4
        for label, value, color in [
            ("Latest Price",   f"{d['latest']} {d['currency']}", "white"),
            ("Change",         f"{d['change']} ({d['pct_change']}%)", "green" if d["pct_change"]>=0 else "red"),
            ("Average Price",  f"{d['avg_price']} {d['currency']}", "white"),
            ("Period High",    f"{d['highest']} {d['currency']}", "green"),
            ("Period Low",     f"{d['lowest']} {d['currency']}", "red"),
            ("Best Day",       f"+{d['best_day']}%", "green"),
            ("Worst Day",      f"{d['worst_day']}%", "red"),
            ("Avg Daily Ret",  f"{d['avg_daily']}%", "white"),
        ]:
            hdr(ws1, r, 1, label); val(ws1, r, 2, value, color); r+=1

        # Indicators
        r = 4
        for label, value, color in [
            ("RSI (14)",        f"{d['rsi']} — {d['rsi_signal']}", "green" if d["rsi_signal"]=="OVERSOLD" else "red" if d["rsi_signal"]=="OVERBOUGHT" else "white"),
            ("MACD Signal",     d["macd_signal"], "green" if "BULL" in str(d["macd_signal"]) else "red"),
            ("Stochastic %K",   f"{d['stoch_k']} — {d['stoch_signal']}", "white"),
            ("Bollinger Band",  d["bb_signal"].replace("_"," "), "white"),
            ("BB Squeeze",      "YES ⚡ Big move!" if d["bb_squeeze"] else "No", "yellow" if d["bb_squeeze"] else "muted"),
            ("ATR Volatility",  f"{d['atr']} ({d['atr_pct']}%)", "white"),
            ("Volume Signal",   f"{d['vol_signal']} ({d['vol_ratio']}x avg)", "white"),
            ("SMA 7",           f"{d['sma7']} {d['currency']}", "white"),
            ("SMA 20",          f"{d['sma20']} {d['currency']}", "white"),
        ]:
            hdr(ws1, r, 3, label); val(ws1, r, 4, value, color); r+=1

        # ── Sheet 2: Price History ──
        ws2 = wb.create_sheet("Price History")
        ws2.sheet_view.showGridLines = False
        for row in ws2.iter_rows(min_row=1, max_row=500, min_col=1, max_col=7):
            for cell in row: cell.fill = dark_fill

        for i, col in enumerate(["DATE","OPEN","HIGH","LOW","CLOSE","VOLUME","CHANGE%"], 1):
            c = ws2.cell(row=1, column=i, value=col)
            c.fill = header_fill
            c.font = Font(color="2979FF", bold=True, name="Consolas")
            ws2.column_dimensions[get_column_letter(i)].width = 16

        for i, row_data in enumerate(d["chart"]["ohlc"], 2):
            ws2.cell(row=i, column=1, value=row_data["t"]).font = Font(color="4A5568", name="Consolas")
            ws2.cell(row=i, column=1).fill = dark_fill
            for j, k in enumerate(["o","h","l","c"], 2):
                c = ws2.cell(row=i, column=j, value=row_data[k])
                c.fill = dark_fill; c.font = Font(color="CDD9E5", name="Consolas")
            vol = d["chart"]["volumes"][i-2] if i-2 < len(d["chart"]["volumes"]) else 0
            ws2.cell(row=i, column=6, value=vol).font = Font(color="CDD9E5", name="Consolas")
            ws2.cell(row=i, column=6).fill = dark_fill
            if i > 2:
                prev_c = d["chart"]["ohlc"][i-3]["c"]
                curr_c = row_data["c"]
                if prev_c and curr_c:
                    chg = round((curr_c - prev_c) / prev_c * 100, 2)
                    c7 = ws2.cell(row=i, column=7, value=chg)
                    c7.fill = dark_fill
                    c7.font = Font(color="00E676" if chg>=0 else "FF1744", name="Consolas")

        # ── Sheet 3: Fundamentals ──
        ws3 = wb.create_sheet("Fundamentals")
        ws3.sheet_view.showGridLines = False
        for row in ws3.iter_rows(min_row=1, max_row=30, min_col=1, max_col=4):
            for cell in row: cell.fill = dark_fill
        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 22
        ws3.column_dimensions["C"].width = 22
        ws3.column_dimensions["D"].width = 22

        f = d["fundamentals"]
        fund_data = [
            ("Company",       f["company"]),
            ("Sector",        f["sector"]),
            ("Industry",      f["industry"]),
            ("Country",       f["country"]),
            ("Employees",     f["employees"]),
            ("Market Cap",    f["market_cap"]),
            ("P/E Ratio",     f["pe_ratio"]),
            ("P/B Ratio",     f["pb_ratio"]),
            ("EPS",           f["eps"]),
            ("Dividend Yield",f["dividend_yield"]),
            ("52W High",      f["week52_high"]),
            ("52W Low",       f["week52_low"]),
            ("Revenue",       f["revenue"]),
            ("Profit Margin", f["profit_margin"]),
            ("Return on Equity",f["roe"]),
            ("Debt/Equity",   f["debt_equity"]),
            ("Current Ratio", f["current_ratio"]),
        ]
        t3 = ws3.cell(row=1, column=1, value="COMPANY FUNDAMENTALS")
        t3.font = Font(color="00E676", bold=True, name="Consolas", size=12)
        t3.fill = dark_fill
        for i, (lbl, vl) in enumerate(fund_data, 3):
            hdr(ws3, i, 1, lbl)
            c = ws3.cell(row=i, column=2, value=str(vl))
            c.fill = dark_fill; c.font = Font(color="CDD9E5", name="Consolas")

        # ── Sheet 4: Support/Resistance ──
        ws4 = wb.create_sheet("Levels")
        ws4.sheet_view.showGridLines = False
        for row in ws4.iter_rows(min_row=1, max_row=30, min_col=1, max_col=4):
            for cell in row: cell.fill = dark_fill
        ws4.column_dimensions["A"].width = 22
        ws4.column_dimensions["B"].width = 18
        ws4.column_dimensions["C"].width = 22
        ws4.column_dimensions["D"].width = 18

        t4 = ws4.cell(row=1, column=1, value="SUPPORT & RESISTANCE + FIBONACCI")
        t4.font = Font(color="00E676", bold=True, name="Consolas", size=12)
        t4.fill = dark_fill

        hdr(ws4, 3, 1, "RESISTANCE LEVELS")
        hdr(ws4, 3, 3, "FIBONACCI LEVELS")
        for i, res in enumerate(d["resistances"], 4):
            c = ws4.cell(row=i, column=1, value="Resistance")
            c.fill = dark_fill; c.font = Font(color="FF1744", name="Consolas")
            c2 = ws4.cell(row=i, column=2, value=res)
            c2.fill = dark_fill; c2.font = Font(color="FF1744", bold=True, name="Consolas")

        curr_row = max(4 + len(d["resistances"]) + 1, 5)
        c = ws4.cell(row=curr_row, column=1, value="CURRENT PRICE")
        c.fill = dark_fill; c.font = Font(color="2979FF", bold=True, name="Consolas")
        c2 = ws4.cell(row=curr_row, column=2, value=d["latest"])
        c2.fill = dark_fill; c2.font = Font(color="2979FF", bold=True, name="Consolas")
        curr_row += 1

        for i, sup in enumerate(d["supports"]):
            c = ws4.cell(row=curr_row+i, column=1, value="Support")
            c.fill = dark_fill; c.font = Font(color="00E676", name="Consolas")
            c2 = ws4.cell(row=curr_row+i, column=2, value=sup)
            c2.fill = dark_fill; c2.font = Font(color="00E676", bold=True, name="Consolas")

        for i, (lbl, fv) in enumerate(d["fib_levels"].items(), 4):
            c = ws4.cell(row=i, column=3, value=lbl)
            c.fill = dark_fill; c.font = Font(color="CE93D8", name="Consolas")
            c2 = ws4.cell(row=i, column=4, value=fv)
            c2.fill = dark_fill; c2.font = Font(color="FFFFFF", bold=True, name="Consolas")

        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{symbol}_analysis.xlsx"
        )

    except ImportError:
        return jsonify({"error":"Please install openpyxl: pip install openpyxl"}), 500
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500



import os as _os

# ── PWA Routes ────────────────────────────────────
@app.route("/manifest.json")
def manifest():
    return send_file("manifest.json", mimetype="application/manifest+json")

@app.route("/sw.js")
def service_worker():
    resp = send_file("sw.js", mimetype="application/javascript")
    resp.headers["Service-Worker-Allowed"] = "/"
    resp.headers["Cache-Control"] = "no-cache"
    return resp

@app.route("/static/<path:filename>")
def static_files(filename):
    return send_file(_os.path.join("static", filename))

# ── AI Assistant Endpoint ─────────────────────────
@app.route("/ai", methods=["POST"])
def ai_chat():
    try:
        import urllib.request
        import json as jsonlib

        body        = request.get_json()
        messages    = body.get("messages", [])
        system      = body.get("system", "")

        payload = jsonlib.dumps({
            "model":      "claude-sonnet-4-20250514",
            "max_tokens": 600,
            "system":     system,
            "messages":   messages[-8:]   # last 8 messages only
        }).encode("utf-8")

        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return jsonify({"error": "ANTHROPIC_API_KEY not set! See instructions below."}), 400

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data    = payload,
            headers = {
                "Content-Type":      "application/json",
                "x-api-key":         api_key,
                "anthropic-version": "2023-06-01"
            },
            method = "POST"
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            data  = jsonlib.loads(resp.read().decode("utf-8"))
            reply = data["content"][0]["text"]
            return jsonify({"reply": reply})

    except urllib.error.HTTPError as e:
        err = e.read().decode("utf-8")
        return jsonify({"error": f"API Error: {err}"}), 500
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    print("=" * 50)
    print("  Stock Analyzer v4.0 starting...")
    print(f"  Open: http://127.0.0.1:{port}")
    print("=" * 50)
    app.run(host="0.0.0.0", port=port, debug=True)
